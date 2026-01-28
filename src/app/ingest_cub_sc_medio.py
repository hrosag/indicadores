#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import sys
import hashlib
from io import BytesIO
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import openpyxl
import pytesseract
from PIL import Image
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout


PAGE_URL = "https://sinduscon-fpolis.org.br/servico/cub-mensal/"
TARGET_SHEET = "2026"  # teste controlado: apenas 2026
OUTPUT_DIR = Path("out")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

MONTHS = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}


@dataclass
class CubMedioRow:
    ano: int
    competencia: str
    competencia_referencia: str
    cub_medio: float
    var_mes: float
    var_ano: float
    var_12m: float
    xlsx_sha256: str
    fonte: str


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def br_to_float(s: str) -> float:
    s = s.strip().replace("%", "")
    s = s.replace(".", "").replace(",", ".")
    return float(s)


def extract_first_image_from_sheet(xlsx_bytes: bytes, sheet_name: str) -> bytes:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba '{sheet_name}' não encontrada. Abas: {wb.sheetnames}")

    ws = wb[sheet_name]
    imgs = getattr(ws, "_images", [])
    if not imgs:
        raise RuntimeError(f"Nenhuma imagem encontrada na aba '{sheet_name}' (ws._images vazio).")

    return imgs[0]._data()


def ocr_image_bytes(img_bytes: bytes) -> str:
    im = Image.open(BytesIO(img_bytes)).convert("L")
    return pytesseract.image_to_string(im, lang="por")


def parse_cub_medio_from_ocr(text: str, sheet_year: int) -> CubMedioRow:
    t = text.upper()

    months = re.findall(r"\b(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)\b", t)
    if len(months) < 2:
        raise RuntimeError(f"Não consegui extrair meses (DEZ/JAN etc.). OCR:\n{text}")

    mes_ref, mes_uso = months[0], months[1]
    use_month_num = MONTHS[mes_uso]
    ref_month_num = MONTHS[mes_ref]

    nums = re.findall(r"\b\d{1,3}(?:\.\d{3})*,\d{2}%?\b", t.replace(" ", ""))
    if len(nums) < 4:
        raise RuntimeError(f"Não consegui extrair 4 números (CUB + 3 %). Achei: {nums}\nOCR:\n{text}")

    cub_medio = br_to_float(nums[0])
    var_mes = br_to_float(nums[1])
    var_ano = br_to_float(nums[2])
    var_12m = br_to_float(nums[3])

    ref_year = sheet_year - 1 if use_month_num == 1 else sheet_year
    competencia = f"{sheet_year:04d}-{use_month_num:02d}"
    competencia_ref = f"{ref_year:04d}-{ref_month_num:02d}"

    return CubMedioRow(
        ano=sheet_year,
        competencia=competencia,
        competencia_referencia=competencia_ref,
        cub_medio=cub_medio,
        var_mes=var_mes,
        var_ano=var_ano,
        var_12m=var_12m,
        xlsx_sha256="",
        fonte="Sinduscon GF - CUB M2 Residencial Médio (planilha com imagem; Playwright + OCR)",
    )


def write_output_xlsx(row: CubMedioRow) -> Path:
    df = pd.DataFrame([{
        "competencia": row.competencia,
        "competencia_referencia": row.competencia_referencia,
        "cub_medio_r$": row.cub_medio,
        "var_mes_%": row.var_mes,
        "var_ano_%": row.var_ano,
        "var_12m_%": row.var_12m,
        "xlsx_sha256": row.xlsx_sha256,
        "fonte": row.fonte,
    }])

    out_path = OUTPUT_DIR / f"cub_sc_residencial_medio_{row.ano}_teste.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=str(row.ano))
    return out_path


def _try_click_text_in_page_or_frames(page, text: str, timeout_ms: int = 90_000) -> None:
    """
    Tenta clicar em texto visível tanto na page quanto em qualquer frame.
    Faz scroll into view e clica.
    """
    # 1) page
    loc = page.get_by_text(text, exact=False).first
    try:
        loc.wait_for(state="visible", timeout=timeout_ms)
        loc.scroll_into_view_if_needed(timeout=timeout_ms)
        loc.click(timeout=timeout_ms)
        return
    except Exception:
        pass

    # 2) frames
    for frame in page.frames:
        try:
            floc = frame.get_by_text(text, exact=False).first
            floc.wait_for(state="visible", timeout=5_000)
            floc.scroll_into_view_if_needed(timeout=5_000)
            floc.click(timeout=5_000)
            return
        except Exception:
            continue

    raise RuntimeError(f"Não consegui localizar/clicar no texto: {text}")


def _dismiss_cookie_banner(page) -> None:
    # Banner típico da página: botão "Ok"
    try:
        ok = page.get_by_text("Ok", exact=False).first
        if ok.is_visible():
            ok.click(timeout=5_000)
    except Exception:
        pass


def download_xlsx_via_clicks() -> bytes:
    """
    Fluxo (confirmado):
      1) 1 - CUB NORMA 2006
      2) 1 - CUB RESIDENCIAL MÉDIO
      3) CUB M2 RESIDENCIAL MÉDIO - ANUAL E MENSAL
      4) Baixar arquivo
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            accept_downloads=True,
            viewport={"width": 1440, "height": 900},
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
        )
        page = context.new_page()
        page.set_default_timeout(90_000)

        try:
            page.goto(PAGE_URL, wait_until="domcontentloaded", timeout=120_000)
            page.wait_for_load_state("networkidle", timeout=120_000)

            _dismiss_cookie_banner(page)

            # Garante que a área “Tabelas CUB” entrou no viewport (ajuda no lazy-load)
            try:
                page.get_by_text("Tabelas CUB", exact=False).first.scroll_into_view_if_needed(timeout=30_000)
            except Exception:
                pass

            # Espera a primeira pasta aparecer em page OU em algum frame
            # (tenta várias vezes rapidamente)
            for _ in range(6):
                try:
                    _try_click_text_in_page_or_frames(page, "1 - CUB NORMA 2006", timeout_ms=15_000)
                    break
                except Exception:
                    page.wait_for_timeout(2_000)
                    _dismiss_cookie_banner(page)
            else:
                raise RuntimeError("A pasta '1 - CUB NORMA 2006' não ficou visível/clicável a tempo.")

            _try_click_text_in_page_or_frames(page, "1 - CUB RESIDENCIAL MÉDIO", timeout_ms=90_000)

            # Item do arquivo (texto visível do cartão)
            _try_click_text_in_page_or_frames(page, "CUB M2 RESIDENCIAL MÉDIO", timeout_ms=90_000)

            # Download (normalmente aparece menu com “Baixar arquivo”)
            with page.expect_download(timeout=90_000) as download_info:
                # às vezes fica dentro de frame, então tenta em ambos
                try:
                    _try_click_text_in_page_or_frames(page, "Baixar arquivo", timeout_ms=30_000)
                except Exception:
                    _try_click_text_in_page_or_frames(page, "Baixar", timeout_ms=30_000)

            download = download_info.value

        except Exception as e:
            # dumps para debug no artifact
            try:
                page.screenshot(path=str(OUTPUT_DIR / "debug_page.png"), full_page=True)
            except Exception:
                pass
            try:
                (OUTPUT_DIR / "debug_page.html").write_text(page.content(), encoding="utf-8")
            except Exception:
                pass
            raise

        tmp_path = OUTPUT_DIR / download.suggested_filename
        download.save_as(str(tmp_path))
        xlsx_bytes = tmp_path.read_bytes()

        context.close()
        browser.close()
        return xlsx_bytes


def main() -> int:
    year = int(TARGET_SHEET)

    xlsx_bytes = download_xlsx_via_clicks()
    xlsx_hash = sha256_bytes(xlsx_bytes)

    img_bytes = extract_first_image_from_sheet(xlsx_bytes, TARGET_SHEET)
    ocr_text = ocr_image_bytes(img_bytes)

    row = parse_cub_medio_from_ocr(ocr_text, sheet_year=year)
    row.xlsx_sha256 = xlsx_hash

    out_path = write_output_xlsx(row)
    (OUTPUT_DIR / f"debug_ocr_{year}.txt").write_text(ocr_text, encoding="utf-8")

    print("OK")
    print(
        f"competencia={row.competencia} cub_medio={row.cub_medio} "
        f"var_mes={row.var_mes} var_ano={row.var_ano} var_12m={row.var_12m}"
    )
    print(f"OUTPUT_XLSX={out_path.as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
